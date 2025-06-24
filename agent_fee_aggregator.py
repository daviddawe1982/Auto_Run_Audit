#!/usr/bin/env python3
"""
Agent Fee Aggregator

This script processes STE_Report Excel files from a network folder structure
and aggregates Agent Fee data by Run and date, outputting results in a
format matching the provided Audit.xlsx example.

Requirements:
- pandas
- openpyxl
"""

import pandas as pd
import os
import sys
from pathlib import Path
from datetime import datetime, timedelta
import re
import requests



from typing import Dict, List, Tuple, Optional

import argparse


class AgentFeeAggregator:
    """Main class for aggregating Agent Fee data from STE_Report files."""
    
    def __init__(self, root_directory: str = r"\\TRUENAS\nasuser\GTS-Data\Reports"):
        """
        Initialize the aggregator.
        
        Args:
            root_directory: Root directory to search for STE_Report files
        """
        self.root_directory = Path(root_directory)
        self.ste_report_files = []
        self.aggregated_data = {}
        
    def find_ste_report_files(self, start_date: Optional[datetime] = None, 
                             end_date: Optional[datetime] = None) -> List[Path]:
        """
        Recursively find all .xlsx files containing 'STE_Report' in filename.
        
        Args:
            start_date: Start date for filtering (optional)
            end_date: End date for filtering (optional)
            
        Returns:
            List of Path objects for matching files
        """
        ste_files = []
        
        if not self.root_directory.exists():
            print(f"Warning: Root directory {self.root_directory} does not exist.")
            return ste_files
            
        # Recursively search for STE_Report*.xlsx files
        for xlsx_file in self.root_directory.rglob("*.xlsx"):
            if "STE_Report" in xlsx_file.name:
                # If date filtering is enabled, try to extract date from path
                if start_date or end_date:
                    file_date = self._extract_date_from_path(xlsx_file)
                    if file_date:
                        if start_date and file_date < start_date:
                            continue
                        if end_date and file_date > end_date:
                            continue
                            
                ste_files.append(xlsx_file)
                
        self.ste_report_files = ste_files
        print(f"Found {len(ste_files)} STE_Report files")
        return ste_files
    
    def _extract_date_from_path(self, file_path: Path) -> Optional[datetime]:
        """
        Extract date from file path based on folder structure.
        Expected structure: ...\\2025\\6 Jun\\20-06-2025\\...
        
        Args:
            file_path: Path object for the file
            
        Returns:
            datetime object if date found, None otherwise
        """
        path_parts = file_path.parts
        
        # Look for date patterns in path
        for part in reversed(path_parts):
            # Check for DD-MM-YYYY pattern
            date_match = re.search(r'(\d{1,2})-(\d{1,2})-(\d{4})', part)
            if date_match:
                try:
                    day, month, year = date_match.groups()
                    return datetime(int(year), int(month), int(day))
                except ValueError:
                    continue
                    
        return None
    
    def process_ste_report_file(self, file_path: Path) -> Dict:
        """
        Process a single STE_Report Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing extracted data
        """
        try:
            # Read the "All Data" worksheet
            df = pd.read_excel(file_path, sheet_name="All Data", engine='openpyxl')
            
            # Check if required columns exist
            if 'Run' not in df.columns or 'Agent Fee' not in df.columns:
                print(f"Warning: Required columns not found in {file_path}")
                print(f"Available columns: {df.columns.tolist()}")
                return {}
            
            # Get file date
            file_date = self._extract_date_from_path(file_path)
            if not file_date:
                print(f"Warning: Could not extract date from {file_path}")
                return {}
            
            # Filter out invalid data (NaN values and zero/negative agent fees)
            df_clean = df.dropna(subset=['Run', 'Agent Fee'])
            df_clean = df_clean[df_clean['Agent Fee'] > 0]
            
            if df_clean.empty:
                print(f"Warning: No valid data found in {file_path}")
                return {}
            
            # Check if Contract column exists for contract-level aggregation
            if 'Contract' in df.columns:
                # Group by Run and Contract, sum Agent Fee
                run_contract_aggregation = df_clean.groupby(['Run', 'Contract'])['Agent Fee'].sum().to_dict()
                
                # Restructure the data
                run_data = {}
                for (run, contract), agent_fee in run_contract_aggregation.items():
                    # Convert run to string for consistency with BEX data
                    run_str = str(run)
                    if run_str not in run_data:
                        run_data[run_str] = {}
                    run_data[run_str][contract] = agent_fee
            else:
                # Fallback: Group by Run only and assume all data is for 'STE' contract
                run_aggregation = df_clean.groupby('Run')['Agent Fee'].sum().to_dict()
                run_data = {}
                for run, agent_fee in run_aggregation.items():
                    # Convert run to string for consistency with BEX data
                    run_str = str(run)
                    run_data[run_str] = {'STE': agent_fee}
            
            return {
                'file_path': file_path,
                'date': file_date,
                'run_data': run_data
            }
            
        except FileNotFoundError:
            print(f"Error: File not found: {file_path}")
            return {}
        except PermissionError:
            print(f"Error: Permission denied accessing: {file_path}")
            return {}
        except ValueError as e:
            if "Worksheet named 'All Data' not found" in str(e):
                print(f"Error: 'All Data' worksheet not found in {file_path}")
            else:
                print(f"Error reading Excel file {file_path}: {e}")
            return {}
        except Exception as e:
            print(f"Unexpected error processing {file_path}: {e}")
            return {}
    
    def aggregate_all_data(self) -> Dict:
        """
        Process all found STE_Report files and aggregate data.
        
        Returns:
            Dictionary with aggregated data by run, contract, and date
        """
        aggregated = {}
        
        for file_path in self.ste_report_files:
            file_data = self.process_ste_report_file(file_path)
            
            if not file_data:
                continue
                
            date = file_data['date']
            # Convert datetime to string format for consistency with BEX data
            date_key = date.strftime("%Y-%m-%d") if isinstance(date, datetime) else str(date)
            run_data = file_data['run_data']
            
            for run, contract_data in run_data.items():
                if run not in aggregated:
                    aggregated[run] = {}
                    
                for contract, agent_fee in contract_data.items():
                    if contract not in aggregated[run]:
                        aggregated[run][contract] = {}
                        
                    if date_key not in aggregated[run][contract]:
                        aggregated[run][contract][date_key] = 0
                        
                    aggregated[run][contract][date_key] += agent_fee
        
        self.aggregated_data = aggregated
        return aggregated
    
    def create_audit_report(self, output_path: str = "Agent_Fee_Audit.xlsx") -> None:
        """
        Create an Excel report in the format of Audit.xlsx with enhanced columns and formulas.
        
        Args:
            output_path: Path for the output Excel file
        """
        if not self.aggregated_data:
            print("No data to export. Run aggregate_all_data() first.")
            return
        
        # Import openpyxl for advanced Excel functionality
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        
        # Prepare data for Excel output
        all_dates = set()
        all_contracts = set()
        
        for run_data in self.aggregated_data.values():
            for contract_data in run_data.values():
                all_dates.update(contract_data.keys())
            all_contracts.update(run_data.keys())
        
        all_dates = sorted(all_dates)
        all_contracts = sorted(all_contracts)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Define cost data structure per run (updated defaults per issue requirements)
        cost_data_template = [
            ('Wage', 0),           # Default to 0 - will be entered manually
            ('Super', 0),          # Default to 0 - will be entered manually  
            ('Running Costs', '=5*30 + 140'),  # Formula: 140 per week + 30 per day
            ('Fuel Liters', 0),    # Default to 0 - will be entered manually
            ('Fuel Cost Per ltr', 0),  # Default to 0 - will be entered manually
            ('Fuel Total', None)   # Will be calculated with formula
        ]
        
        # Different values for different runs (using new defaults per issue requirements)
        run_specific_costs = {
            20: [
                ('Wage', 0),              # Default to 0 - will be entered manually
                ('Super', 0),             # Default to 0 - will be entered manually
                ('Running Costs', '=5*30 + 140'),  # Formula: 140 per week + 30 per day
                ('Fuel Liters', 0),       # Default to 0 - will be entered manually  
                ('Fuel Cost Per ltr', 0), # Default to 0 - will be entered manually
                ('Fuel Total', None)
            ],
            32: [
                ('Wage', 0),              # Default to 0 - will be entered manually
                ('Super', 0),             # Default to 0 - will be entered manually
                ('Running Costs', '=5*30 + 140'),  # Formula: 140 per week + 30 per day
                ('Fuel Liters', 0),       # Default to 0 - will be entered manually
                ('Fuel Cost Per ltr', 0), # Default to 0 - will be entered manually
                ('Fuel Total', None)
            ]
        }
        
        current_row = 1
        
        for run in sorted(self.aggregated_data.keys(), key=lambda x: int(str(x)) if str(x).isdigit() else float('inf')):
            # Get cost data for this run
            cost_data = run_specific_costs.get(run, cost_data_template)
            
            # Add Run header in A1
            ws.cell(row=current_row, column=1, value=f"Run {run} Audit")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            
            current_row += 1
            
            # Add date range in A2 if dates exist
            if all_dates:
                start_date_str = min(all_dates) if all_dates else ""
                end_date_str = max(all_dates) if all_dates else ""
                date_range = f"({start_date_str} to {end_date_str})" if start_date_str and end_date_str else ""
                date_cell = ws.cell(row=current_row, column=1, value=date_range)
                date_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Merge B1 and B2 and center them
                ws.merge_cells(f'A{current_row}:B{current_row}')
                merged_cell = ws.cell(row=current_row-1, column=2)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
            
            # Add headers row - use day names instead of dates
            headers_row = current_row
            
            # Set row height to 32 (convert to points: 32 * 0.75 = 24 points)
            ws.row_dimensions[headers_row].height = 32
            
            # Create alignment for headers (center horizontal and vertical, with text wrap)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Add headers with formatting
            header_cell = ws.cell(row=headers_row, column=1, value="Contract Name")
            header_cell.font = Font(bold=True)
            header_cell.alignment = header_alignment
            
            # Day name columns (B to F) - MON, TUE, WED, THUR, FRI
            day_names = ["MON", "TUE", "WED", "THUR", "FRI"]
            for i, day_name in enumerate(day_names, 2):
                header_cell = ws.cell(row=headers_row, column=i, value=day_name)
                header_cell.font = Font(bold=True)
                header_cell.alignment = header_alignment
            
            # Updated column positions (removing empty columns I, K, L, P, R)
            header_cell = ws.cell(row=headers_row, column=7, value="Totals")
            header_cell.font = Font(bold=True)
            header_cell.alignment = header_alignment
            
            header_cell = ws.cell(row=headers_row, column=8, value="Revenue Day Rate")
            header_cell.font = Font(bold=True)
            header_cell.alignment = header_alignment
            
            header_cell = ws.cell(row=headers_row, column=9, value="Week Total")
            header_cell.font = Font(bold=True)
            header_cell.alignment = header_alignment
            
            header_cell = ws.cell(row=headers_row, column=10, value="Cost")
            header_cell.font = Font(bold=True)
            header_cell.alignment = header_alignment
            
            header_cell = ws.cell(row=headers_row, column=12, value="Cost Day Rate")
            header_cell.font = Font(bold=True)
            header_cell.alignment = header_alignment
            
            current_row += 1
            contract_start_row = current_row
            
            # Add contract data for this run
            run_contracts = self.aggregated_data[run]
            contract_rows = []
            
            for contract in sorted(all_contracts):
                if contract in run_contracts:
                    contract_rows.append(current_row)
                    # Contract name
                    ws.cell(row=current_row, column=1, value=contract)
                    
                    # Agent fee data for each date (limit to 5 dates)
                    for col_idx, date in enumerate(all_dates[:5], 2):
                        agent_fee = run_contracts[contract].get(date, 0)
                        if agent_fee > 0:
                            ws.cell(row=current_row, column=col_idx, value=agent_fee)
                    
                    # Column G: Totals (SUM of daily values B to F) - was column H
                    ws.cell(row=current_row, column=7, value=f"=SUM(B{current_row}:F{current_row})")
                    
                    current_row += 1
            
            # Add empty rows for cost structure - we need exactly 6 more rows with SUM formulas
            for i in range(6):
                ws.cell(row=current_row, column=7, value=f"=SUM(B{current_row}:F{current_row})")  # was column 8
                current_row += 1
            
            # Now add the cost data and formulas
            # The main contract row (first actual contract with data)
            if contract_rows:
                main_contract_row = contract_rows[0]
                cost_start_row = main_contract_row
                cost_end_row = current_row - 1
                
                # Column H: Revenue Day Rate = Week Total / 5 (only for first contract) - was column J
                ws.cell(row=main_contract_row, column=8, value=f"=I{main_contract_row}/5")
                
                # Column I: Week Total = SUM of all G values in this section - was column M
                ws.cell(row=main_contract_row, column=9, value=f"=SUM(G{cost_start_row}:G{cost_end_row})")
                
                # Add cost breakdown in columns J and K - was columns N and O
                for i, (cost_item, cost_value) in enumerate(cost_data):
                    cost_row = main_contract_row + i
                    ws.cell(row=cost_row, column=10, value=cost_item)  # Column J (was N)
                    
                    if cost_item == "Fuel Total":
                        # Fuel Total = Fuel Cost Per ltr * Fuel Liters
                        fuel_liters_row = main_contract_row + 3  # Fuel Liters row
                        fuel_cost_row = main_contract_row + 4    # Fuel Cost Per ltr row
                        ws.cell(row=cost_row, column=11, value=f"=K{fuel_cost_row}*K{fuel_liters_row}")  # was O
                    elif isinstance(cost_value, str) and cost_value.startswith('='):
                        ws.cell(row=cost_row, column=11, value=cost_value)  # Column K (was O)
                    elif cost_value is not None:
                        ws.cell(row=cost_row, column=11, value=cost_value)  # Column K (was O)
                
                # Column L: Cost Day Rate = (sum of all costs) / 5 - was column Q
                ws.cell(row=main_contract_row, column=12, value=f"=SUM(K{main_contract_row}:K{main_contract_row + 5}) / 5")
                
                # Add Factor header and formula in column L, positioned after cost items
                factor_row = main_contract_row + 1  # Skip one row after first cost item
                factor_header_cell = ws.cell(row=factor_row, column=12, value="Factor")
                factor_header_cell.font = Font(bold=True)
                factor_header_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                factor_formula_cell = ws.cell(row=factor_row + 1, column=12, value=f"=H{main_contract_row}/L{main_contract_row}")
                
                # Add Revenue header and formula in column L, positioned after Factor
                revenue_row = main_contract_row + 3  # Two rows after Factor
                revenue_header_cell = ws.cell(row=revenue_row, column=12, value="Revenue")
                revenue_header_cell.font = Font(bold=True)
                revenue_header_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                revenue_formula_cell = ws.cell(row=revenue_row + 1, column=12, value=f"=I{main_contract_row}-SUM(K{main_contract_row}:K{main_contract_row + 5})")
            
            # Add separation rows
            current_row += 2
        
        # Apply enhanced styling and formatting
        from openpyxl.styles import PatternFill, Border, Side
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        light_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        medium_border = Border(left=Side(style='medium'), right=Side(style='medium'), 
                             top=Side(style='medium'), bottom=Side(style='medium'))
        
        # Track run sections for border application
        run_sections = []
        current_section = None
        
        # Identify run sections - don't apply any default borders yet
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                
                # Check if this is a run header
                if cell_value and isinstance(cell_value, str) and "Audit" in cell_value:
                    if current_section:
                        run_sections.append(current_section)
                    current_section = {'start_row': row_idx, 'start_col': 1}
        
        # Close the last section
        if current_section:
            current_section['end_row'] = ws.max_row
            current_section['end_col'] = 12  # Column L is the rightmost data column (Cost Day Rate, Factor, Revenue)
            run_sections.append(current_section)
        
        # Apply styling to headers and cells - but NO automatic borders yet
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Bold formatting and background for main headers and audit titles
                    if ("Audit" in str(cell.value) or 
                        cell.value in ["Contract Name", "MON", "TUE", "WED", "THUR", "FRI", 
                                      "Totals", "Revenue Day Rate", "Week Total", 
                                      "Cost", "Cost Day Rate"]):
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = header_fill
                    elif cell.value in ["Wage", "Super", "Running Costs", "Fuel Liters", 
                                       "Fuel Cost Per ltr", "Fuel Total"]:
                        cell.font = Font(bold=True)
                        cell.fill = light_fill
                    elif cell.value in ["Factor", "Revenue"]:
                        # Factor and Revenue headers should have the same blue background and white text as other headers
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = header_fill
                elif cell.value and isinstance(cell.value, (int, float)):
                    # Number formatting for numeric values
                    cell.number_format = '0.00'
        
        # Apply borders exactly matching the example.xlsx pattern
        for section in run_sections:
            start_row = section['start_row']
            end_row = section.get('end_row', ws.max_row)
            start_col = section['start_col']
            end_col = section.get('end_col', 12)
            
            # Find the actual end row by looking for the next run or end of data
            actual_end_row = start_row
            for check_row in range(start_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row=check_row, column=1).value
                if cell_value and isinstance(cell_value, str) and "Audit" in cell_value:
                    actual_end_row = check_row - 2  # Stop before the next run (with gap)
                    break
                # Check if there's any content in this row
                has_content = any(ws.cell(row=check_row, column=col).value is not None 
                                for col in range(1, end_col + 1))
                if has_content:
                    actual_end_row = check_row

            # Apply borders matching the exact example.xlsx pattern
            for row_idx in range(start_row, actual_end_row + 1):
                for col_idx in range(start_col, end_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # Determine relative position within the section
                    rel_row = row_idx - start_row + 1  # 1-based relative row
                    col_letter = chr(64 + col_idx)
                    
                    # Define the exact border pattern based on example.xlsx analysis
                    # This maps (relative_row, column_letter) to border configuration
                    border_map = {
                        # Row 1 patterns
                        (1, 'A'): ('medium', 'thin', 'medium', 'thin'),      # L:medium|R:thin|T:medium|B:thin
                        (1, 'B'): ('thin', None, 'medium', None),            # L:thin|T:medium
                        (1, 'C'): (None, None, 'medium', None),              # T:medium
                        (1, 'D'): (None, None, 'medium', None),              # T:medium
                        (1, 'E'): (None, None, 'medium', None),              # T:medium
                        (1, 'F'): (None, None, 'medium', None),              # T:medium
                        (1, 'G'): (None, None, 'medium', None),              # T:medium
                        (1, 'H'): (None, None, 'medium', None),              # T:medium
                        (1, 'I'): (None, None, 'medium', None),              # T:medium
                        (1, 'J'): (None, None, 'medium', None),              # T:medium
                        (1, 'K'): (None, None, 'medium', None),              # T:medium
                        (1, 'L'): (None, 'medium', 'medium', None),          # R:medium|T:medium
                        
                        # Row 2 patterns
                        (2, 'A'): ('medium', None, None, 'thin'),            # L:medium|B:thin
                        (2, 'B'): (None, None, None, 'thin'),                # B:thin
                        (2, 'C'): (None, None, None, 'thin'),                # B:thin
                        (2, 'D'): (None, None, None, 'thin'),                # B:thin
                        (2, 'E'): (None, None, None, 'thin'),                # B:thin
                        (2, 'F'): (None, None, None, 'thin'),                # B:thin
                        (2, 'G'): (None, None, None, 'thin'),                # B:thin
                        (2, 'H'): (None, None, None, 'thin'),                # B:thin
                        (2, 'I'): (None, None, None, 'thin'),                # B:thin
                        (2, 'J'): (None, None, None, 'thin'),                # B:thin
                        (2, 'K'): (None, None, None, None),                  # (no value in example)
                        (2, 'L'): (None, 'medium', None, 'thin'),            # R:medium|B:thin
                        
                        # Row 3 patterns (headers)
                        (3, 'A'): ('medium', 'thin', 'thin', 'thin'),        # L:medium|R:thin|T:thin|B:thin
                        (3, 'B'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (3, 'C'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (3, 'D'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (3, 'E'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (3, 'F'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (3, 'G'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (3, 'H'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (3, 'I'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (3, 'J'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (3, 'K'): ('thin', 'thin', None, 'thin'),            # L:thin|R:thin|B:thin
                        (3, 'L'): ('thin', 'medium', 'thin', 'thin'),        # L:thin|R:medium|T:thin|B:thin
                        
                        # Row 4 patterns (main data row)
                        (4, 'A'): ('medium', 'thin', 'thin', 'thin'),        # L:medium|R:thin|T:thin|B:thin
                        (4, 'B'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (4, 'C'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (4, 'D'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (4, 'E'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (4, 'F'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (4, 'G'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (4, 'H'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (4, 'I'): ('thin', None, 'thin', 'thin'),            # L:thin|T:thin|B:thin
                        (4, 'J'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (4, 'K'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (4, 'L'): ('thin', 'medium', 'thin', 'thin'),        # L:thin|R:medium|T:thin|B:thin
                        
                        # Row 5 patterns
                        (5, 'A'): ('medium', 'thin', 'thin', 'thin'),        # L:medium|R:thin|T:thin|B:thin
                        (5, 'B'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (5, 'C'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (5, 'D'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (5, 'E'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (5, 'F'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (5, 'G'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (5, 'H'): ('thin', None, 'thin', None),              # L:thin|T:thin
                        (5, 'I'): (None, None, None, None),                  # (no borders)
                        (5, 'J'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (5, 'K'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (5, 'L'): ('thin', 'medium', 'thin', 'thin'),        # L:thin|R:medium|T:thin|B:thin
                        
                        # Row 6 patterns
                        (6, 'A'): ('medium', 'thin', 'thin', 'thin'),        # L:medium|R:thin|T:thin|B:thin
                        (6, 'B'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (6, 'C'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (6, 'D'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (6, 'E'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (6, 'F'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (6, 'G'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (6, 'H'): ('thin', None, None, None),                # L:thin
                        (6, 'I'): (None, None, None, None),                  # (no borders)
                        (6, 'J'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (6, 'K'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (6, 'L'): ('thin', 'medium', 'thin', None),          # L:thin|R:medium|T:thin
                        
                        # Row 7 patterns
                        (7, 'A'): ('medium', 'thin', 'thin', 'thin'),        # L:medium|R:thin|T:thin|B:thin
                        (7, 'B'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (7, 'C'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (7, 'D'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (7, 'E'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (7, 'F'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (7, 'G'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (7, 'H'): ('thin', None, None, None),                # L:thin
                        (7, 'I'): (None, None, None, None),                  # (no borders)
                        (7, 'J'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (7, 'K'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (7, 'L'): ('thin', 'medium', 'thin', 'thin'),        # L:thin|R:medium|T:thin|B:thin
                        
                        # Row 8 patterns
                        (8, 'A'): ('medium', 'thin', 'thin', 'thin'),        # L:medium|R:thin|T:thin|B:thin
                        (8, 'B'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (8, 'C'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (8, 'D'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (8, 'E'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (8, 'F'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (8, 'G'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (8, 'H'): ('thin', None, None, None),                # L:thin
                        (8, 'I'): (None, None, None, None),                  # (no borders)
                        (8, 'J'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (8, 'K'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (8, 'L'): ('thin', 'medium', 'thin', 'thin'),        # L:thin|R:medium|T:thin|B:thin
                        
                        # Row 9 patterns
                        (9, 'A'): ('medium', 'thin', 'thin', 'thin'),        # L:medium|R:thin|T:thin|B:thin
                        (9, 'B'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (9, 'C'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (9, 'D'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (9, 'E'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (9, 'F'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (9, 'G'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (9, 'H'): ('thin', None, None, None),                # L:thin
                        (9, 'I'): (None, None, None, None),                  # (no borders)
                        (9, 'J'): ('thin', 'thin', 'thin', 'thin'),          # L:thin|R:thin|T:thin|B:thin
                        (9, 'K'): ('thin', 'thin', 'thin', None),            # L:thin|R:thin|T:thin
                        (9, 'L'): ('thin', 'medium', None, None),            # L:thin|R:medium
                        
                        # Row 10 patterns
                        (10, 'A'): ('medium', 'thin', 'thin', 'thin'),       # L:medium|R:thin|T:thin|B:thin
                        (10, 'B'): ('thin', 'thin', 'thin', 'thin'),         # L:thin|R:thin|T:thin|B:thin
                        (10, 'C'): ('thin', 'thin', 'thin', 'thin'),         # L:thin|R:thin|T:thin|B:thin
                        (10, 'D'): ('thin', 'thin', 'thin', 'thin'),         # L:thin|R:thin|T:thin|B:thin
                        (10, 'E'): ('thin', 'thin', 'thin', 'thin'),         # L:thin|R:thin|T:thin|B:thin
                        (10, 'F'): ('thin', 'thin', 'thin', 'thin'),         # L:thin|R:thin|T:thin|B:thin
                        (10, 'G'): ('thin', 'thin', 'thin', 'thin'),         # L:thin|R:thin|T:thin|B:thin
                        (10, 'H'): ('thin', None, None, None),               # L:thin
                        (10, 'I'): (None, None, None, None),                 # (no borders)
                        (10, 'J'): (None, None, None, None),                 # (no borders)
                        (10, 'K'): (None, None, 'thin', None),               # T:thin
                        (10, 'L'): (None, 'medium', None, None),             # R:medium
                        
                        # Row 11 patterns (bottom border)
                        (11, 'A'): ('medium', None, 'thin', 'medium'),       # L:medium|T:thin|B:medium
                        (11, 'B'): (None, None, 'thin', 'medium'),           # T:thin|B:medium
                        (11, 'C'): (None, None, 'thin', 'medium'),           # T:thin|B:medium
                        (11, 'D'): (None, None, 'thin', 'medium'),           # T:thin|B:medium
                        (11, 'E'): (None, None, 'thin', 'medium'),           # T:thin|B:medium
                        (11, 'F'): (None, None, 'thin', 'medium'),           # T:thin|B:medium
                        (11, 'G'): (None, None, 'thin', 'medium'),           # T:thin|B:medium
                        (11, 'H'): (None, None, None, 'medium'),             # B:medium
                        (11, 'I'): (None, None, None, 'medium'),             # B:medium
                        (11, 'J'): (None, None, None, 'medium'),             # B:medium
                        (11, 'K'): (None, None, None, 'medium'),             # B:medium
                        (11, 'L'): (None, 'medium', None, 'medium'),         # R:medium|B:medium
                    }
                    
                    # Get border configuration for this cell
                    border_config = border_map.get((rel_row, col_letter), ('thin', 'thin', 'thin', 'thin'))
                    left_style, right_style, top_style, bottom_style = border_config
                    
                    # Apply the border
                    cell.border = Border(
                        left=Side(style=left_style) if left_style else Side(),
                        right=Side(style=right_style) if right_style else Side(),
                        top=Side(style=top_style) if top_style else Side(),
                        bottom=Side(style=bottom_style) if bottom_style else Side()
                    )
        
        # Set specific column widths as per requirements
        # B,C,D,E,F,G should be 10 wide
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 10
        
        # A,H,I,J,K,L should be 14 wide  
        for col_letter in ['A', 'H', 'I', 'J', 'K', 'L']:
            ws.column_dimensions[col_letter].width = 14
        
        # Save the workbook
        wb.save(output_path)
        print(f"Enhanced audit report saved to {output_path}")


def fetch_bex_contract_data(session: requests.Session, start_date: datetime, end_date: datetime) -> Dict:
    """
    Fetch BEX contract data from TransVirtual portal for runs 1-50.
    
    Args:
        session: Authenticated requests session
        start_date: Start date for data range
        end_date: End date for data range
        
    Returns:
        Dictionary with BEX contract data aggregated by run and date
    """
    bex_data = {}
    date_range_str = f"{start_date.strftime('%b %d, %Y')}%20-%20{end_date.strftime('%b %d, %Y')}" if start_date and end_date else ""
    
    # Check runs 1 to 50
    for run_num in range(1, 51):
        run = f"run{run_num:02d}"
        print(f"Fetching data for {run}...")
        
        # Construct URL with dynamic date range
        manifest_url = (
            f"https://portal.transvirtual.com/Portal/Consignment/ManifestListGrid?_search=true&nd=1750731459270&rows=100&page=1&sidx=&sord=desc"
            f"&dateRange={date_range_str}&idSearchQueries=0&manifestType=Runsheet&createdBy.UserFirstName={run}"
        )
        
        try:
            r = session.get(manifest_url)
            response_data = r.json()
            
            if not response_data.get('rows'):
                print(f"  No data found for {run}")
                continue
                
            run_daily_totals = {}  # date -> total amount
            row_count = len(response_data['rows'])
            print(f"  Found {row_count} manifests for {run}")
            
            for row in response_data['rows']:
                id_number = row['id']
                manifest_date = row['cell'][2]  # Date in format DD/MM/YYYY
                
                # Convert date format from DD/MM/YYYY to YYYY-MM-DD for consistency
                try:
                    date_obj = datetime.strptime(manifest_date, "%d/%m/%Y")
                    date_key = date_obj.strftime("%Y-%m-%d")
                except ValueError:
                    print(f"  Invalid date format: {manifest_date}")
                    continue
                
                # Get detailed data for this manifest
                detail_url = (
                    f"https://portal.transvirtual.com/Portal/Consignment/ManifestDetailGrid?_search=false"
                    f"&nd=&rows=500&page=1&sidx=&sord=desc&idManifest={id_number}&tbleName=ManifestRunsheetDetail"
                )
                
                res = session.get(detail_url)
                res_data = res.json()
                
                # Extract the ConsignmentCustomerBaseTotal
                try:
                    total = float(res_data['userdata']['ConsignmentCustomerBaseTotal'])
                    
                    # Aggregate by date
                    if date_key not in run_daily_totals:
                        run_daily_totals[date_key] = 0
                    run_daily_totals[date_key] += total
                    
                except (KeyError, ValueError, TypeError) as e:
                    print(f"  Error processing manifest {id_number}: {e}")
                    continue
            
            # Only add run to bex_data if it has actual data
            if run_daily_totals:
                bex_data[str(run_num)] = {'BEX': run_daily_totals}
                print(f"  {run}: Found data for {len(run_daily_totals)} dates, total: {sum(run_daily_totals.values()):.2f}")
            else:
                print(f"  {run}: No valid data found")
                
        except Exception as e:
            print(f"  Error fetching data for {run}: {e}")
            continue
    
    return bex_data


def get_date_range() -> Tuple[Optional[datetime], Optional[datetime]]:
    """
    Get date range from user input.
    
    Returns:
        Tuple of (start_date, end_date) or (None, None) if no filtering
    """
    print("\nDate Range Selection:")
    print("1. Process all files (no date filtering)")
    print("2. Specify date range")
    
    choice = input("Enter choice (1 or 2): ").strip()
    
    if choice == "1":
        return None, None
    elif choice == "2":
        try:
            start_str = input("Enter start date (YYYY-MM-DD): ").strip()
            end_str = input("Enter end date (YYYY-MM-DD): ").strip()
            
            start_date = datetime.strptime(start_str, "%Y-%m-%d") if start_str else None
            end_date = datetime.strptime(end_str, "%Y-%m-%d") if end_str else None
            
            return start_date, end_date
        except ValueError as e:
            print(f"Invalid date format: {e}")
            return None, None
    else:
        print("Invalid choice. Processing all files.")
        return None, None

# TransVirtual Login
payload = {
        'UserNameLogin': 'David@gee',
        'PasswordLogin': 'khhed06!!!',
        'RememberMe': 'false'
    }
headers = {
        "Host": "portal.transvirtual.com",
        "Connection": "keep-alive",
        "Content-Length": "67",
        "Origin": "https://portal.transvirtual.com",
        "X-Requested-With": "XMLHttpRequest",
        "Content-Type": "application/json",
        "Accept": "*/*",
        "Referer": "https://portal.transvirtual.com/Public/Home/Login?ReturnUrl=%2FPortal%2FConsignment%2FManifestListGrid%3F_search%3Dfalse%26nd%3D1650803811508%26rows%3D1000%26page%3D1%26sidx%3D%26sord%3Ddesc%26dateRange%3DMar%2B26%252C%2B2022%2B-%2BApr%2B24%252C%2B2022%26idSearchQueries%3D0%26manifestType%3DRunsheet",
        "Accept-Encoding": "gzip,deflate,br",
        "Accept-Language": "en-AU,en-GB;q=0.9,en-US;q=0.8,en;q=0.7",
        "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
        "pragma": "no-cache",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "Windows",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.199 Safari/537.36",
    }


def main():
    """Main function to run the Agent Fee Aggregator."""
    parser = argparse.ArgumentParser(description="Aggregate Agent Fee data from STE_Report files")
    parser.add_argument("--root-dir", default=r"\\TRUENAS\nasuser\GTS-Data\Reports",
                       help="Root directory to search for STE_Report files")
    parser.add_argument("--output", default="Agent_Fee_Audit.xlsx",
                       help="Output Excel file path")
    parser.add_argument("--start-date", help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end-date", help="End date (YYYY-MM-DD)")
    parser.add_argument("--interactive", action="store_true",
                       help="Run in interactive mode for date selection")
    
    args = parser.parse_args()
    
    # Initialize aggregator
    aggregator = AgentFeeAggregator(args.root_dir)
    
    # Get date range
    if args.interactive:
        start_date, end_date = get_date_range()
    else:
        start_date = datetime.strptime(args.start_date, "%Y-%m-%d") if args.start_date else None
        end_date = datetime.strptime(args.end_date, "%Y-%m-%d") if args.end_date else None
    
    print(f"\nSearching for STE_Report files in: {aggregator.root_directory}")
    if start_date or end_date:
        print(f"Date range: {start_date} to {end_date}")
    
    # Find and process STE files
    files = aggregator.find_ste_report_files(start_date, end_date)
    
    print(f"Found {len(files)} STE_Report files")
    if files:
        print(f"\nProcessing {len(files)} STE files...")
        aggregated_data = aggregator.aggregate_all_data()
    else:
        print("No STE_Report files found.")
        aggregated_data = {}
    
    # Initialize empty aggregated_data if no STE files found
    if not aggregated_data:
        print("No valid STE data found in the files.")
        aggregated_data = {}

    print(f"\nFound STE data for {len(aggregated_data)} runs")
    for run in sorted(aggregated_data.keys(), key=lambda x: int(str(x)) if str(x).isdigit() else float('inf')):
        contracts = list(aggregated_data[run].keys())
        print(f"  Run {run}: {len(contracts)} contracts ({', '.join(contracts)})")
        for contract in contracts:
            dates_count = len(aggregated_data[run][contract])
            print(f"    {contract}: {dates_count} dates")
    
    # Fetch BEX contract data from TransVirtual
    s = requests.Session()
    def TV_Login():
        try:
            response = s.post('https://portal.transvirtual.com/Public/Home/LoginCommit', data=payload, headers=headers)
            return response.status_code == 200
        except Exception as e:
            print(f"Login failed: {e}")
            return False
    
    print("\nLogging in to TV to fetch BEX contract data...")    
    if not TV_Login():
        print("Failed to login to TransVirtual. Skipping BEX data fetch.")
        bex_data = {}
    else:
        # Fetch BEX contract data for all runs
        bex_data = fetch_bex_contract_data(s, start_date, end_date)
    
    # Integrate BEX data with existing aggregated_data
    print(f"\nIntegrating BEX data...")
    for run_key, run_contracts in bex_data.items():
        # run_key is already a string from BEX data fetch
        
        if run_key not in aggregated_data:
            aggregated_data[run_key] = {}
            
        # Add BEX contract data
        for contract, dates_data in run_contracts.items():
            if contract not in aggregated_data[run_key]:
                aggregated_data[run_key][contract] = {}
                
            for date, amount in dates_data.items():
                if date not in aggregated_data[run_key][contract]:
                    aggregated_data[run_key][contract][date] = 0
                aggregated_data[run_key][contract][date] += amount
    
    # Update aggregator's data
    aggregator.aggregated_data = aggregated_data
    
    # Check if we have any data at all (STE or BEX)
    if not aggregated_data:
        print("No valid data found (neither STE nor BEX data).")
        return
    
    # Print final summary
    print(f"\nFinal aggregated data summary:")
    for run in sorted(aggregated_data.keys(), key=lambda x: int(str(x)) if str(x).isdigit() else float('inf')):
        contracts = list(aggregated_data[run].keys())
        print(f"  Run {run}: {len(contracts)} contracts ({', '.join(contracts)})")
        for contract in contracts:
            dates_count = len(aggregated_data[run][contract])
            total_amount = sum(aggregated_data[run][contract].values())
            print(f"    {contract}: {dates_count} dates, total: {total_amount:.2f}")
        
    # Create output report
    print(f"\nCreating audit report: {args.output}")
    aggregator.create_audit_report(args.output)
    
    print("Process completed successfully!")


if __name__ == "__main__":
    main()
